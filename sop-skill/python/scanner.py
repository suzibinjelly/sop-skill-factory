#!/usr/bin/env python3
"""scanner.py - Recursively scans a target directory and parses files into
structured JSON output for the SOP Skill Factory.

CLI:
    python scanner.py --output <output_json_path> --target <target_directory>

Exit codes: 0=success, 1=business error, 2=JSON error, 3=IO/permission error
"""

import argparse
import csv
import io
import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

# ---------------------------------------------------------------------------
# Budget constants
# ---------------------------------------------------------------------------
MAX_TOTAL_CHARS = 50000
MAX_SINGLE_FILE_CHARS = 50000
LARGE_FILE_CHARS = 10000
PREVIEW_CHARS = 500

# ---------------------------------------------------------------------------
# Exclusion rules
# ---------------------------------------------------------------------------
EXCLUDED_DIRS = {".git", "node_modules", "__pycache__", ".sop-temp", ".sop-skill-cache"}

TEXT_EXTENSIONS = {
    ".md", ".txt", ".yaml", ".yml", ".json", ".csv", ".html", ".htm",
}

SPECIAL_EXTENSIONS = {
    ".docx": "docx",
    ".pdf": "pdf",
    ".xlsx": "xlsx",
}

ALL_SUPPORTED_EXTENSIONS = TEXT_EXTENSIONS | set(SPECIAL_EXTENSIONS.keys())

TYPE_LABELS = {
    ".md": "markdown",
    ".txt": "text",
    ".yaml": "yaml",
    ".yml": "yaml",
    ".json": "json",
    ".csv": "csv",
    ".html": "html",
    ".htm": "html",
    ".docx": "docx",
    ".pdf": "pdf",
    ".xlsx": "xlsx",
}

# ---------------------------------------------------------------------------
# Encoding detection helpers
# ---------------------------------------------------------------------------

def detect_encoding(raw_bytes: bytes) -> Optional[str]:
    """Try charset-normalizer first, then chardet."""
    try:
        from charset_normalizer import from_bytes as cn_from_bytes
        results = cn_from_bytes(raw_bytes)
        if results and results.best():
            return results.best().encoding
    except Exception:
        pass

    try:
        import chardet
        result = chardet.detect(raw_bytes)
        if result and result.get("encoding"):
            return result["encoding"]
    except Exception:
        pass

    return None


def decode_bytes(raw_bytes: bytes) -> tuple[Optional[str], Optional[str]]:
    """Decode bytes to string. Returns (text, encoding) or (None, None)."""
    # Try UTF-8 first (most common case)
    try:
        return raw_bytes.decode("utf-8"), "utf-8"
    except (UnicodeDecodeError, ValueError):
        pass

    # Detect encoding
    enc = detect_encoding(raw_bytes)
    if enc:
        try:
            return raw_bytes.decode(enc), enc
        except (UnicodeDecodeError, LookupError, ValueError):
            pass

    return None, None

# ---------------------------------------------------------------------------
# File-type readers
# ---------------------------------------------------------------------------

def read_text_file(file_path: Path) -> tuple[Optional[str], Optional[str], Optional[str]]:
    """Read a plain-text file. Returns (text, encoding, error)."""
    try:
        raw = file_path.read_bytes()
    except (IOError, PermissionError) as exc:
        return None, None, str(exc)

    text, enc = decode_bytes(raw)
    if text is None:
        return None, None, "Could not detect file encoding"
    return text, enc, None


def read_docx(file_path: Path) -> tuple[Optional[str], Optional[str], Optional[str]]:
    """Read a .docx file via python-docx. Returns (text, encoding, error)."""
    try:
        from docx import Document
    except ImportError:
        return None, None, "python-docx is not installed"
    try:
        doc = Document(str(file_path))
        paragraphs = [p.text for p in doc.paragraphs]
        return "\n".join(paragraphs), None, None
    except Exception as exc:
        return None, None, str(exc)


def read_pdf(file_path: Path) -> tuple[Optional[str], Optional[str], Optional[str]]:
    """Read a .pdf file via pymupdf/fitz. Returns (text, encoding, error)."""
    try:
        import fitz
    except ImportError:
        return None, None, "pymupdf is not installed"
    try:
        doc = fitz.open(str(file_path))
        pages = [page.get_text() for page in doc]
        doc.close()
        return "\n".join(pages), None, None
    except Exception as exc:
        return None, None, str(exc)


def read_xlsx(file_path: Path) -> tuple[Optional[str], Optional[str], Optional[str]]:
    """Read first sheet of .xlsx as CSV-like text. Returns (text, encoding, error)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None, None, "openpyxl is not installed"
    try:
        wb = load_workbook(str(file_path), read_only=True, data_only=True)
        ws = wb.active
        buf = io.StringIO()
        writer = csv.writer(buf)
        for row in ws.iter_rows(values_only=True):
            writer.writerow([str(c) if c is not None else "" for c in row])
        wb.close()
        return buf.getvalue(), None, None
    except Exception as exc:
        return None, None, str(exc)

# ---------------------------------------------------------------------------
# Reader dispatcher
# ---------------------------------------------------------------------------

def read_file(file_path: Path) -> tuple[Optional[str], Optional[str], Optional[str]]:
    """Dispatch to the appropriate reader based on extension.
    Returns (text, encoding, error).
    """
    ext = file_path.suffix.lower()
    if ext in SPECIAL_EXTENSIONS:
        readers = {
            "docx": read_docx,
            "pdf": read_pdf,
            "xlsx": read_xlsx,
        }
        return readers[SPECIAL_EXTENSIONS[ext]](file_path)
    return read_text_file(file_path)

# ---------------------------------------------------------------------------
# Title extraction
# ---------------------------------------------------------------------------

def extract_title(text: Optional[str], file_path: Path) -> str:
    """Extract a meaningful title from file content or fall back to filename."""
    if not text:
        return file_path.stem

    ext = file_path.suffix.lower()

    # Markdown: first # heading
    if ext == ".md":
        for line in text.splitlines():
            stripped = line.strip()
            if stripped.startswith("#"):
                title = stripped.lstrip("#").strip()
                if title:
                    return title

    # HTML: first <title> or <h1>
    if ext in (".html", ".htm"):
        m = re.search(r"<title[^>]*>(.*?)</title>", text, re.IGNORECASE | re.DOTALL)
        if m and m.group(1).strip():
            return m.group(1).strip()
        m = re.search(r"<h1[^>]*>(.*?)</h1>", text, re.IGNORECASE | re.DOTALL)
        if m and m.group(1).strip():
            # Strip inner tags
            return re.sub(r"<[^>]+>", "", m.group(1)).strip()

    return file_path.stem

# ---------------------------------------------------------------------------
# Scanning logic
# ---------------------------------------------------------------------------

def should_skip(entry: os.DirEntry) -> bool:
    """Return True if the entry should be excluded."""
    name = entry.name
    # Hidden files / directories
    if name.startswith("."):
        return True
    # Excluded directories
    if entry.is_dir(follow_symlinks=False) and name in EXCLUDED_DIRS:
        return True
    return False


def is_binary(file_path: Path) -> bool:
    """Heuristic binary detection: check for null bytes in first 8 KB."""
    try:
        with open(file_path, "rb") as f:
            chunk = f.read(8192)
        return b"\x00" in chunk
    except (IOError, PermissionError):
        return True


def collect_files(target_dir: Path) -> list[Path]:
    """Walk the target directory and return a sorted list of supported files."""
    collected: list[Path] = []

    def _walk(directory: Path):
        try:
            entries = list(os.scandir(directory))
        except (PermissionError, OSError):
            return
        for entry in sorted(entries, key=lambda e: e.name):
            if should_skip(entry):
                continue
            entry_path = Path(entry.path)
            if entry.is_dir(follow_symlinks=False):
                _walk(entry_path)
            elif entry.is_file(follow_symlinks=False):
                ext = entry_path.suffix.lower()
                if ext not in ALL_SUPPORTED_EXTENSIONS:
                    continue
                if is_binary(entry_path) and ext in TEXT_EXTENSIONS:
                    continue
                collected.append(entry_path)

    _walk(target_dir)
    return collected


def process_file(file_path: Path, target_dir: Path) -> dict:
    """Process a single file into its output dict."""
    rel = file_path.relative_to(target_dir)
    rel_posix = rel.as_posix()
    ext = file_path.suffix.lower()
    file_type = TYPE_LABELS.get(ext, "unknown")

    result: dict = {
        "path": rel_posix,
        "type": file_type,
        "encoding": None,
        "char_count": 0,
        "title": file_path.stem,
        "preview": None,
        "full_text": None,
        "error": None,
    }

    text, encoding, error = read_file(file_path)

    if error:
        result["error"] = error
        result["encoding"] = encoding
        return result

    if text is None:
        result["error"] = "No text extracted"
        return result

    result["encoding"] = encoding
    result["char_count"] = len(text)
    result["title"] = extract_title(text, file_path)
    result["preview"] = text[:PREVIEW_CHARS]
    result["full_text"] = text  # may be set to None later by budget logic
    return result

# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def scan(target_dir: Path, output_path: Path) -> int:
    """Run the scan and write JSON output. Returns exit code."""
    if not target_dir.is_dir():
        print(f"Error: target is not a directory: {target_dir}", file=sys.stderr)
        return 3

    files = collect_files(target_dir)

    # First pass: process every file
    file_results: list[dict] = []
    for fp in files:
        file_results.append(process_file(fp, target_dir))

    # Calculate total characters (only non-oversized files)
    total_chars = sum(r["char_count"] for r in file_results if r["error"] is None)

    # Determine budget strategy
    full_text_available = True

    # Mark oversized files (single file > MAX_SINGLE_FILE_CHARS)
    for r in file_results:
        if r["char_count"] > MAX_SINGLE_FILE_CHARS:
            r["full_text"] = None
            r["error"] = "File exceeds maximum single-file size limit"
            total_chars -= r["char_count"]
            r["char_count"] = 0
            r["preview"] = None

    # If total exceeds budget, downgrade all files to preview-only
    if total_chars > MAX_TOTAL_CHARS:
        full_text_available = False
        for r in file_results:
            if r["error"] is None:
                r["full_text"] = None
    else:
        # For large individual files, keep only preview
        for r in file_results:
            if r["error"] is None and r["char_count"] > LARGE_FILE_CHARS:
                r["full_text"] = None

    # Recompute total_chars accurately
    total_chars = sum(r["char_count"] for r in file_results)

    output = {
        "meta": {
            "scan_time": datetime.now(timezone.utc).isoformat(),
            "target_dir": str(target_dir.resolve()),
            "total_files": len(file_results),
            "total_chars": total_chars,
            "full_text_available": full_text_available,
        },
        "files": file_results,
    }

    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(output, f, ensure_ascii=False, indent=2)
    except (TypeError, ValueError) as exc:
        print(f"JSON serialization error: {exc}", file=sys.stderr)
        return 2
    except (IOError, PermissionError) as exc:
        print(f"IO error writing output: {exc}", file=sys.stderr)
        return 3

    print(f"Scan complete: {len(file_results)} files, {total_chars} chars -> {output_path}")
    return 0


def main():
    parser = argparse.ArgumentParser(
        description="Scan a directory and produce structured JSON for SOP Skill Factory"
    )
    parser.add_argument(
        "--output",
        required=True,
        help="Path to the output JSON file",
    )
    parser.add_argument(
        "--target",
        required=True,
        help="Target directory to scan",
    )
    args = parser.parse_args()

    target_dir = Path(args.target)
    output_path = Path(args.output)

    try:
        code = scan(target_dir, output_path)
    except Exception as exc:
        print(f"Unexpected error: {exc}", file=sys.stderr)
        sys.exit(1)

    sys.exit(code)


if __name__ == "__main__":
    main()
